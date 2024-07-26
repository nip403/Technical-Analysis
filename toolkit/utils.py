import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from ta import add_all_ta_features
from openpyxl import load_workbook
import yfinance as yf
from functools import wraps
import random

figsize = (15, 9)
candle_interval = 3 # biweekly

def _clean_xl(file: str) -> None:
    wb = load_workbook(file)
    
    if "Test" in wb.sheetnames:
        del wb["Test"]
        
    wb.save(file)
    
def write(file: str, data: pd.DataFrame, sheet_name: str = "Test") -> None:
    try: # in case file is open
        _clean_xl(file)
        
        with pd.ExcelWriter(file, mode="a") as fh:
            data.to_excel(fh, sheet_name=sheet_name)
    except:
        return
    
def rand_hex() -> str:
    return f"#{''.join(f'{random.randint(0, 255):02X}' for _ in range(3))}"

def norm(vec: list[int | float]) -> list[float]:
    vals = np.fromiter(vec, dtype=float)
    return vals / np.sum(vals)

def process_data_ticker(ticker: str, features: bool = True) -> pd.DataFrame: # yfinance api is the worst thing ever invented 
    ticker = yf.Ticker(ticker).history(period="max").tail(1000)[["Open", "High", "Low", "Close", "Volume"]] # limit to 4 years
    
    ticker.index = ticker.index.date
    ticker.reset_index(inplace=True)
    ticker.columns = ["Date", "Open", "High", "Low", "Close", "Volume"]
    
    return ticker.dropna() if not features else add_all_ta_features(
        ticker.dropna(),
        open="Open",
        high="High",
        low="Low",
        close="Close",
        volume="Volume",
    )

def process_data(file_name: str, sheet_name: str, features: bool = True) -> pd.DataFrame:
    data = pd.read_excel(file_name, sheet_name=sheet_name)
    
    assert set("Open High Low Close Volume".split() if features else ["Close"]).issubset(data.columns)
    
    return data.dropna() if not features else add_all_ta_features(
        data.dropna(),
        open="Open",
        high="High",
        low="Low",
        close="Close",
        volume="Volume",
    )
    
def process_ticker_multiple(tickers: str, features: bool = True):
    return {
        i: process_data_ticker(i, features) for i in tickers
    }
    
def process_data_multiple(file_name: str, sheet_names: list[str], features: bool = True) -> list[pd.DataFrame]:
    return {
        i: process_data(file_name, i, features) for i in sheet_names
    }

def _to_pct(vals: dict) -> dict:
    return {
        k: round(100 * v, 2) for k, v in vals.items()
    }
    
class AxisHandler: # TODO
    def __init__(self, axes: dict, **kwargs):
        self.axes = {int(k[2:]): v for k, v in axes.items() if "ax" in k} # requires format ax[number]
        
        for k, v in self.axes.items():
            setattr(self, f"ax{k}", v)
            
    def __getitem__(self, key):
        return self.axes[key]

    def __setitem__(self, key, value):
        self.axes[key] = value
        
def plotter(simple=True, rows=2, ratios=None, plot_candlestick=False, plot_volume=False, rescale=[]):
    """Plotting wrapper
    
    Decorates a function(
        data: output of utils.process_data()
        *,
        candlestick: bool - plot stock price using candlesticks
        volume: bool - plot stock trading volumes on the same subplot
    )
        => function can access all specified axes (ax1 ... axn) and data
        
    Usage:
        @plotter()
        def graph(data, **kwargs):
            return
            
    Note: 
        price data will always be plotted on ax1, to remove, use ax1.remove() and the .change_geometry methods accordingly

    Args:
        simple (bool, optional): 
            Plots a simple graph: Price vs Date; False if specifying 2 or more rows. 
            Defaults to True.
        rows (int, optional): 
            Specifies the number of different plots needed. 
            Defaults to 2.
        ratios (_type_, optional):
            Ratios of the sizes of each graph(/row) in order top to bottom. 
            e.g. for 3 rows, [1, 2, 3], bottom graph is 3x taller than top graph
            Defaults to None.
        plot_candlestick (bool, optional):
            Plots candlestick price movements.
            Groups data into units of 3 (=candle_interval) dates for each candle. 
            Defaults to False.
        plot_volume (bool, optional): 
            Plot volume on price graph. 
            Defaults to False.
        rescale (list, optional):
            Since axis.set_xlim() is apparently irreversible, this is an option to request a list of axis indices (0-based) to exclude from x_lim setting.
            e.g. [0, 2] excludes ax1, ax3 from xlim.
            Defaults to [].
    """
    
    if ratios is not None:
        assert rows == len(ratios)

    def dec(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            candlestick = kwargs.get("candlestick", plot_candlestick)
            volume = kwargs.get("volume", plot_volume)
            _figsize = kwargs.get("figsize", figsize)
            df = args[0]
            
            if simple:
                fig, ax1 = plt.subplots(figsize=figsize)
                kwargs["ax1"] = ax1
                axes = [ax1]
            else: 
                fig, axes = plt.subplots(rows, 1, figsize=_figsize, gridspec_kw={"height_ratios": [4, 1] if ratios is None else ratios})
                ax1 = axes[0]
                
                for p, axis in enumerate(axes):
                    kwargs[f"ax{p + 1}"] = axis
                    axis.set_xlabel("Date")
                    axis.set_xlim(df["Date"].iloc[0], df["Date"].iloc[-1]) # strips empty space
            
            ax1.set_xlabel("Date")
            ax1.set_ylabel("Price")
            
            func(*args, **kwargs)
            
            # grid shenanigans
            ax1.grid(True)
            ax1.minorticks_on()
            ax1.grid(which="major", color="#DDDDDD", linewidth=0.8)
            ax1.grid(which="minor", color="#DDDDDD", linewidth=0.5, axis="y")
            
            if not candlestick:
                ax1.plot(df["Date"], df["Close"], label="Price", color="black")
            else:
                df2 = df.groupby(df.index // candle_interval).agg({ # dont show every day
                    "Date": "first",
                    "Close": "mean",
                    "Open": "mean",
                    "Low": "min",
                    "High": "max",
                })
                
                df2.reset_index(drop=True, inplace=True)
                up = df2[df2["Close"] >= df2["Open"]]
                down = df2[df2["Close"] < df2["Open"]]
                
                ax1.bar(up["Date"], up["Close"] - up["Open"], 4, bottom=up["Open"], color="green")
                ax1.bar(up["Date"], up["High"] - up["Close"], 2, bottom=up["Close"], color="green")
                ax1.bar(up["Date"], up["Low"] - up["Open"], 2, bottom=up["Open"], color="green")

                ax1.bar(down["Date"], down["Close"] - down["Open"], 4, bottom=down["Open"], color="red")
                ax1.bar(down["Date"], down["High"] - down["Open"], 2, bottom=down["Open"], color="red")
                ax1.bar(down["Date"], down["Low"] - down["Close"], 2, bottom=down["Close"], color="red")

                ax1.grid(False, which="minor")
            
            # label latest price
            latest = df.iloc[-1]
            ax1.annotate(
                f"<{round(latest['Close'], 2)}>",
                xy=(1.01, latest["Close"]),
                xycoords=("axes fraction", "data"),
                horizontalalignment="left", 
                verticalalignment="center",
                color="red",
            )
            
            # plot volume
            if volume:
                vol = ax1.twinx()
                vol.bar(df["Date"], df["Volume"], color="blue", alpha=0.5)
                vol.set_ylabel("Volume", color="blue")
                vol.tick_params(axis="y", labelcolor="blue")
                vol.yaxis.set_ticks_position("left") 
                vol.yaxis.set_label_position("left")
            
            for p, a in enumerate(axes):
                a.legend(loc="upper left")
                a.yaxis.set_ticks_position("right") 
                a.yaxis.set_label_position("right")
                
                if not p in rescale:
                    a.set_xlim(df["Date"].iloc[0], df["Date"].iloc[-1])
            
            ax1.set_ylim(bottom=0)
            fig.autofmt_xdate()
            fig.tight_layout()
            
            if kwargs.get("DEBUG_1", False):
                fig.subplots_adjust(hspace=0.1, wspace=0.1)
            
            plt.show()
        return wrapper
    return dec
